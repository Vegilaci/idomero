from io import BytesIO
import re

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from app import models
from app.database import get_db


router = APIRouter(prefix="/api/export", tags=["Export"])

INVALID_SHEET_CHARACTERS = re.compile(r"[\\/*?:\[\]]")


def format_time(time_ms: int | None) -> str:
    if time_ms is None:
        return "-"

    hours, remainder = divmod(time_ms, 3_600_000)
    minutes, remainder = divmod(remainder, 60_000)
    seconds, milliseconds = divmod(remainder, 1_000)

    return f"{hours:02}:{minutes:02}:{seconds:02}.{milliseconds:03}"


def create_unique_sheet_name(name: str, member_id: int, used_names: set[str]) -> str:
    cleaned_name = INVALID_SHEET_CHARACTERS.sub("_", name or "").strip()
    cleaned_name = cleaned_name or f"Versenyzo_{member_id}"
    base_name = cleaned_name[:31]
    sheet_name = base_name
    counter = 2

    while sheet_name.casefold() in used_names:
        suffix = f" ({counter})"
        sheet_name = f"{base_name[:31 - len(suffix)]}{suffix}"
        counter += 1

    used_names.add(sheet_name.casefold())
    return sheet_name


def set_column_widths(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=0,
        )
        column_letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 3, 12),
            36,
        )


@router.get("/members/excel")
def export_members_excel(db: Session = Depends(get_db)):
    members = (
        db.query(models.Member)
        .options(
            joinedload(models.Member.team),
            joinedload(models.Member.laps),
        )
        .order_by(models.Member.name.asc(), models.Member.id.asc())
        .all()
    )

    workbook = Workbook()
    workbook.remove(workbook.active)

    title_font = Font(size=16, bold=True)
    label_font = Font(bold=True)
    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    best_lap_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
    used_sheet_names: set[str] = set()

    if not members:
        worksheet = workbook.create_sheet("Nincs adat")
        worksheet["A1"] = "Nincs exportĂˇlhatĂł versenyzĹ‘."
        worksheet["A1"].font = title_font
    else:
        for member in members:
            worksheet = workbook.create_sheet(
                create_unique_sheet_name(member.name, member.id, used_sheet_names)
            )

            laps = sorted(
                member.laps,
                key=lambda lap: (
                    lap.lap_no if lap.lap_no is not None else 0,
                    lap.id,
                ),
            )
            lap_times = [lap.time_ms for lap in laps if lap.time_ms is not None]
            total_time_ms = sum(lap_times)
            best_lap_ms = min(lap_times) if lap_times else None
            average_lap_ms = (
                round(total_time_ms / len(lap_times)) if lap_times else None
            )
            team_name = member.team.name if member.team else "Nincs csapat"

            worksheet["A1"] = member.name
            worksheet["A1"].font = title_font

            summary_rows = [
                ("RajtszĂˇm", member.rajt_szam),
                ("Csapat", team_name),
                ("MĂ©rt kĂ¶rĂ¶k", len(laps)),
                ("Ă–sszidĹ‘", format_time(total_time_ms)),
                ("Ătlag kĂ¶ridĹ‘", format_time(average_lap_ms)),
                ("Legjobb kĂ¶r", format_time(best_lap_ms)),
            ]

            for row_index, (label, value) in enumerate(summary_rows, start=3):
                worksheet.cell(row=row_index, column=1, value=label).font = label_font
                worksheet.cell(row=row_index, column=2, value=value)

            header_row = 11
            headers = [
                "KĂ¶r",
                "KĂ¶ridĹ‘",
                "KĂ¶ridĹ‘ (ms)",
                "EltĂ©rĂ©s a legjobb kĂ¶rtĹ‘l",
            ]

            for column_index, header in enumerate(headers, start=1):
                cell = worksheet.cell(
                    row=header_row,
                    column=column_index,
                    value=header,
                )
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row_index, lap in enumerate(laps, start=header_row + 1):
                difference_ms = (
                    lap.time_ms - best_lap_ms
                    if lap.time_ms is not None and best_lap_ms is not None
                    else None
                )

                worksheet.cell(row=row_index, column=1, value=lap.lap_no)
                worksheet.cell(row=row_index, column=2, value=format_time(lap.time_ms))
                worksheet.cell(row=row_index, column=3, value=lap.time_ms)
                worksheet.cell(
                    row=row_index,
                    column=4,
                    value=(
                        f"+{format_time(difference_ms)}"
                        if difference_ms is not None
                        else "-"
                    ),
                )

                if lap.time_ms == best_lap_ms:
                    for column_index in range(1, 5):
                        worksheet.cell(
                            row=row_index,
                            column=column_index,
                        ).fill = best_lap_fill

            worksheet.freeze_panes = "A12"
            worksheet.auto_filter.ref = (
                f"A11:D{max(header_row, header_row + len(laps))}"
            )
            set_column_widths(worksheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": 'attachment; filename="versenyzok_export.xlsx"'
        },
    )
